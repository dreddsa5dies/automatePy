#! python
# blankRowInserter.py - вставка строк в файлы excel.

import sys, logging, openpyxl, os
from openpyxl.cell import column_index_from_string

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')

# Usage
if len(sys.argv) == 1:
    print('''Usage: py.exe blankRowInserter.py <int_1> <int_2> <file>
    int_1 - с какой строки вставка
    int_2 - сколько строк вставить
    file - имя файла Excel''')

elif len(sys.argv) == 4:
    # проверка вх данных
    val_1, val_2, fileName = sys.argv[1], sys.argv[2], sys.argv[3]
    if val_1.isdecimal() and val_2.isdecimal and os.path.isfile(fileName) and fileName.endswith('.xlsx'):
        val_1 = int(val_1)
        val_2 = int(val_2)
        fileName = os.path.abspath(fileName)
        logging.info('Введены значения %i %i %s' % (val_1, val_2, fileName))

        wb = openpyxl.load_workbook(fileName)
        sheet = wb.get_active_sheet()
        
        maxRow = sheet.get_highest_row()
        logging.info('Всего строк %s' % (maxRow))
        if maxRow < val_1:
            logging.info('Строк и так меньше, выход...')
            sys.exit()
        
        tmp = {}
        # получение и добавление строк с сохранением во временном значении
        for i in range(maxRow):
            if i < val_1-1:
                tmp[i] = sheet.rows[i]
            elif i == val_1-1:
                for y in range(val_2):
                    tmp[i+y] = ()
                tmp[i+y+1] = sheet.rows[i]
            else:
                tmp[i+y+1] = sheet.rows[i]
        
        wbTmp = openpyxl.Workbook()
        sheetTmp = wbTmp.get_active_sheet()
        for v in tmp:
            for obj in tmp[v]:
                # v - это теперь номер строки
                #  obj.coordinate надо сделать новым адресом
                sheetTmp.cell(row=v+1, column=column_index_from_string(obj.column)).value = obj.value
        
        wbTmp.save('insert_multiplication.xlsx')
        logging.info('Выход...')
    else:
        print('Usage: py.exe blankRowInserter.py <int_1> <int_2> <file>')