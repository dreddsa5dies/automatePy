#! python
# invertTable.py - инвертирование таблиц в файле excel.

import openpyxl, sys

if len(sys.argv) == 1 or len(sys.argv) > 2:
	print('Usage: py.exe invertTable.py <file>')

elif len(sys.argv) == 2:
	# проверка вх данных
	fileName = sys.argv[1]
	if fileName.endswith('.xlsx'):
		wb = openpyxl.load_workbook(fileName)
		sheet = wb.get_active_sheet()
		maxRow = sheet.get_highest_row()
		maxCol = sheet.get_highest_column()

		wbTmp = openpyxl.Workbook()
		sheetTmp = wbTmp.get_active_sheet()

		for i in range(maxRow):
			for k in range(maxCol):
				sheetTmp.cell(row=k+1, column=i+1).value = sheet.cell(row=i+1, column=k+1).value
		
		wbTmp.save('invert_table.xlsx')